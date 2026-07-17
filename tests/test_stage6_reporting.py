import json
from pathlib import Path
import tempfile
import unittest
import zipfile

import pandas as pd
from openpyxl import load_workbook

from src.excel_reporting import HEADER_FILL,SHEET_ORDER,create_excel_workbook
from src.file_loader import load_config
from src.package_outputs import build_reproducibility_metadata,create_analysis_bundle,create_file_manifest,save_reproducibility,sha256_file
from src.reporting import build_cross_test_comparison,build_pressure_derived_summary,create_markdown_reports


def context(tmp):
    encoder=pd.DataFrame([{"Run_ID":"run_001","Nominal_Target_Cycle_s":2.5,"Target_Source":"recorded","Final_Selected_Period_s":2.51,"Final_Selected_Frequency_Hz":1/2.51,"Timing_Method_Confidence":.8,"Valid_Interval_Count":4,"Valid_Sample_CV_Percent":.3}])
    vfd=pd.DataFrame([{"Run_ID":"run_001","Desired_Target_Frequency_Hz":None,"Reconstructed_Command_mV_Capped":None,"Command_Equivalent_Cycle_s":None,"Command_Saturation_Flag":False,"VFD_Verification_Status":"unavailable_unverified_scaling"}])
    pressure=pd.DataFrame([{"Run_ID":"run_001","Robust_Upstream_P5_P95_Span":10.,"Robust_Downstream_P5_P95_Span":3.,"Robust_Attenuation_Ratio":.3,"Median_Cycle_Attenuation_Ratio":.31,"Dynamic_Turbine_DeltaP_RMS":4.,"Dynamic_Turbine_DeltaP_P5_P95_Span":8.}])
    torque=pd.DataFrame([{"Run_ID":"run_001","Mean_Torque":5.,"Peak_Torque":8.,"RMS_Torque":6.}]); gen=pd.DataFrame([{"Run_ID":"run_001","Mean":3.2,"Linear_Drift_Per_s":.01,"Drift_R_Squared":.1,"Combined_GenV_Classification":"drifting_nonperiodic"}]); quality=pd.DataFrame([{"Run_ID":"run_001","Signal":"Pressure_4","Check":"pressure_pair_offset","Severity":"review","Evidence":"offset"}])
    tables={"encoder_summary":encoder,"vfd":vfd,"pressure_response":pressure,"torque":torque,"generator":gen,"quality_findings":quality}
    cross=build_cross_test_comparison(tables)
    for key in ["intake_audit","boundaries","steady","encoder_methods","descriptive","pressure_pairs","pressure_phase","pressure_consistency","torque_phase","correlation","cycle_level","encoder_intervals"]: tables[key]=pd.DataFrame({"Status":["available"]})
    metadata={"input_filename":"synthetic.csv","input_sha256":"a"*64,"configuration_sha256":"b"*64,"analysis_timestamp":"2026-01-01T00:00:00Z","configuration_profile":"new_turbine_2_to_3_second","worksheet_analyzed":None,"total_records":100,"available_channels":["TimeStamp","Encoder","Torque"],"missing_expected_channels":["Pressure_1"],"vfd_scaling_status":"unavailable_unverified_scaling","output_directory":str(tmp),"git_commit":"test","python_version":"3.12"}
    audit={"timestamp_start":"2026-01-01","timestamp_end":"2026-01-02","sampling_interval_median_s":.012,"median_derived_sampling_rate_hz":83.333,"effective_mean_sampling_rate_hz":82.5,"missing_value_count":0,"duplicate_row_count":0}
    return {"tables":tables,"cross_test":cross,"metadata":metadata,"audit":audit}


class Stage6ReportingTests(unittest.TestCase):
    def setUp(self): self.temp=tempfile.TemporaryDirectory(); self.root=Path(self.temp.name); self.config=load_config("config/new_turbine_config.yaml"); self.ctx=context(self.root)
    def tearDown(self): self.temp.cleanup()
    def workbook(self):
        path=create_excel_workbook(self.root/"analysis_summary.xlsx",self.ctx,self.config,pd.DataFrame({"Relative_Path":["x.csv"]})); return path,load_workbook(path)

    def test_cross_test_preserves_decimal_target(self): self.assertEqual(build_cross_test_comparison(self.ctx["tables"]).Target_Cycle_s.iloc[0],2.5)
    def test_cross_test_source_label(self): self.assertEqual(self.ctx["cross_test"].Target_Source.iloc[0],"recorded")
    def test_cross_test_primary_quality(self): self.assertIn("Pressure_4",self.ctx["cross_test"].Primary_Quality_Finding.iloc[0])
    def test_unknown_pressure_units(self): self.assertEqual(self.config["sensor_units"]["Pressure_1"],"unknown")
    def test_derived_summary_unknown_units(self): self.assertTrue((build_pressure_derived_summary(pd.DataFrame(),self.config["sensor_units"]).Units=="unknown").all())
    def test_excel_creation(self): path,wb=self.workbook(); self.assertTrue(path.is_file()); wb.close()
    def test_required_sheet_names_and_order(self): _,wb=self.workbook(); self.assertEqual(wb.sheetnames,SHEET_ORDER); wb.close()
    def test_header_formatting(self): _,wb=self.workbook(); self.assertEqual(wb["Cross_Test_Comparison"]["A1"].fill.fgColor.rgb[-6:],HEADER_FILL.fgColor.rgb[-6:]); wb.close()
    def test_frozen_panes(self): _,wb=self.workbook(); self.assertEqual(wb["Cross_Test_Comparison"].freeze_panes,"A2"); wb.close()
    def test_auto_filter(self): _,wb=self.workbook(); self.assertIsNotNone(wb["Cross_Test_Comparison"].auto_filter.ref); wb.close()
    def test_numeric_formatting(self): _,wb=self.workbook(); self.assertNotEqual(wb["Cross_Test_Comparison"]["B2"].number_format,"General"); wb.close()
    def test_warning_conditional_formatting(self): _,wb=self.workbook(); self.assertGreater(len(wb["Signal_Quality"].conditional_formatting),0); wb.close()
    def test_missing_section_documented(self): self.ctx["tables"]["torque"]=pd.DataFrame(); _,wb=self.workbook(); self.assertIn("unavailable",str(wb["Torque_Analysis"]["B1"].value)); wb.close()
    def test_unavailable_style(self): _,wb=self.workbook(); self.assertTrue(any(c.fill.fgColor.rgb for row in wb["VFD_Verification"].iter_rows() for c in row)); wb.close()
    def test_markdown_reports_created(self): self.assertEqual(len(create_markdown_reports(self.root,self.ctx,self.config)),4)
    def test_engineering_report_sections(self): create_markdown_reports(self.root,self.ctx,self.config); text=(self.root/"report/engineering_analysis_report.md").read_text(); self.assertIn("## 13. Conclusion",text)
    def test_methods_names_software(self): create_markdown_reports(self.root,self.ctx,self.config); self.assertIn("openpyxl",(self.root/"report/methods_section.md").read_text())
    def test_executive_summary_length(self): create_markdown_reports(self.root,self.ctx,self.config); n=len((self.root/"report/executive_summary.md").read_text().split()); self.assertGreater(n,250); self.assertLess(n,550)
    def test_limitations_document(self): create_markdown_reports(self.root,self.ctx,self.config); self.assertIn("New-turbine configuration",(self.root/"report/limitations_and_recommendations.md").read_text())
    def test_report_display_precision(self): create_markdown_reports(self.root,self.ctx,self.config); self.assertIn("0.012",(self.root/"report/engineering_analysis_report.md").read_text())
    def test_weak_phase_wording(self): create_markdown_reports(self.root,self.ctx,self.config); self.assertIn("Weak-correlation phase",(self.root/"report/engineering_analysis_report.md").read_text())
    def test_genv_low_r2_wording(self): create_markdown_reports(self.root,self.ctx,self.config); self.assertIn("low R²",(self.root/"report/engineering_analysis_report.md").read_text())
    def test_four_point_regression_caution(self): self.ctx["cross_test"]=pd.concat([self.ctx["cross_test"]]*4,ignore_index=True); create_markdown_reports(self.root,self.ctx,self.config); self.assertIn("n = 4",(self.root/"report/engineering_analysis_report.md").read_text())
    def test_new_turbine_vfd_unavailable(self): create_markdown_reports(self.root,self.ctx,self.config); self.assertIn("no constants were invented",(self.root/"report/engineering_analysis_report.md").read_text())
    def test_sha256(self): p=self.root/"x"; p.write_text("abc"); self.assertEqual(sha256_file(p),"ba7816bf8f01cfea414140de5dae2223b00361a396177a9cb410ff61f20015ad")
    def test_reproducibility_metadata_hashes(self):
        inp=self.root/"input.csv"; inp.write_text("a,b\n1,2\n"); cfg=self.root/"config.yaml"; cfg.write_text("column_aliases: {}\n")
        meta=build_reproducibility_metadata(inp,cfg,self.config,self.root,None,1,["a","b"],["test"]); self.assertEqual(meta["input_sha256"],sha256_file(inp)); self.assertEqual(meta["configuration_sha256"],sha256_file(cfg))
    def test_reproducibility_files(self): paths=save_reproducibility(self.root,self.ctx["metadata"],self.config); self.assertTrue(all(p.is_file() for p in paths)); self.assertEqual(json.loads(paths[0].read_text())["input_filename"],"synthetic.csv")
    def test_file_manifest(self): (self.root/"a.csv").write_text("a\n1\n"); path,frame=create_file_manifest(self.root); self.assertTrue(path.is_file()); self.assertIn("a.csv",frame.Relative_Path.tolist())
    def test_zip_contents_and_input_exclusion(self):
        (self.root/"analysis_summary.xlsx").write_text("x"); (self.root/"file_manifest.txt").write_text("x"); (self.root/"reproducibility_metadata.json").write_text("{}"); (self.root/"cleaned").mkdir(); (self.root/"cleaned/a.csv").write_text("x")
        bundle=create_analysis_bundle(self.root); self.assertTrue(zipfile.is_zipfile(bundle)); names=zipfile.ZipFile(bundle).namelist(); self.assertIn("cleaned/a.csv",names); self.assertNotIn("input.xlsx",names)
    def test_profiles_reporting_compatible(self): self.assertTrue(load_config("config/legacy_config.yaml")["reporting"]["enabled"]); self.assertTrue(load_config("config/new_turbine_config.yaml")["reporting"]["enabled"])
    def test_repository_synthetic_fixture_targets(self):
        fixture=pd.read_csv("sample_data/synthetic_new_turbine_stage6.csv"); self.assertEqual(sorted(fixture.Target_Cycle_s.unique().tolist()),[2.0,2.5,3.0]); self.assertTrue(pd.api.types.is_float_dtype(fixture.Target_Cycle_s))
    def test_repository_synthetic_fixture_uses_unverified_vfd_profile(self): self.assertFalse(load_config("config/new_turbine_config.yaml")["vfd_command"]["enabled"])


if __name__=="__main__": unittest.main()
