"""Microsoft Excel engineering-summary workbook generation using openpyxl."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook,load_workbook
from openpyxl.formatting.rule import CellIsRule,FormulaRule
from openpyxl.styles import Alignment,Font,PatternFill
from openpyxl.utils import get_column_letter


SHEET_ORDER=["Overview","Intake_Audit","Run_Boundaries","Steady_State","VFD_Verification","Encoder_Cycle_Summary","Encoder_Methods","Descriptive_Statistics","Cross_Test_Comparison","Pressure_Response","Pressure_Relationships","Pressure_Phase","Pressure_Consistency","Torque_Analysis","Torque_Phase","Generator_Voltage","Correlation_Regression","Signal_Quality","Cycle_Level_Data","Methods","Engineering_Interpretation","Limitations","Recommended_Tests","File_Manifest"]
HEADER_FILL=PatternFill("solid",fgColor="1F4E78"); HEADER_FONT=Font(color="FFFFFF",bold=True); WARNING_FILL=PatternFill("solid",fgColor="FCE4D6"); INFERRED_FILL=PatternFill("solid",fgColor="FFF2CC"); RECONSTRUCTED_FILL=PatternFill("solid",fgColor="DDEBF7"); UNAVAILABLE_FILL=PatternFill("solid",fgColor="E7E6E6")


def _excel_value(value:Any)->Any:
    if value is pd.NA or (isinstance(value,(float,np.floating)) and np.isnan(value)): return None
    if isinstance(value,np.generic): return value.item()
    if isinstance(value,pd.Timestamp): return value.to_pydatetime()
    return value


def _format_table(ws,frame:pd.DataFrame,freeze:bool=True,auto_filter:bool=True)->None:
    if frame.empty:
        ws.append(["Status","unavailable — required analysis or channel was not present"]); frame=pd.DataFrame()
    else:
        ws.append(list(frame.columns))
        for row in frame.itertuples(index=False,name=None): ws.append([_excel_value(v) for v in row])
    for cell in ws[1]: cell.fill=HEADER_FILL; cell.font=HEADER_FONT; cell.alignment=Alignment(wrap_text=True,vertical="top")
    if freeze: ws.freeze_panes="A2"
    if auto_filter and ws.max_row>1: ws.auto_filter.ref=ws.dimensions
    for col in range(1,ws.max_column+1):
        header=str(ws.cell(1,col).value or ""); width=min(42,max(10,len(header)+2,*[min(40,len(str(ws.cell(r,col).value or ""))) for r in range(2,min(ws.max_row,150)+1)])); ws.column_dimensions[get_column_letter(col)].width=width
        for row in range(2,ws.max_row+1):
            cell=ws.cell(row,col); cell.alignment=Alignment(wrap_text=True,vertical="top")
            if isinstance(cell.value,(int,float)):
                if "Percent" in header: cell.number_format="0.000"
                elif "Count" in header or header.endswith("Samples"): cell.number_format="0"
                elif "Command" in header or "mV" in header: cell.number_format="0.00"
                elif "Gen" in header: cell.number_format="0.0000"
                else: cell.number_format="0.000"
            elif hasattr(cell.value,"year"): cell.number_format="yyyy-mm-dd hh:mm:ss.000"
    if ws.max_row>1:
        ws.conditional_formatting.add(ws.dimensions,FormulaRule(formula=[f'OR(ISNUMBER(SEARCH("warning",A2)),ISNUMBER(SEARCH("review",A2)))'],fill=WARNING_FILL))
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                text=str(cell.value or "").lower()
                if "inferred" in text: cell.fill=INFERRED_FILL
                elif "reconstructed" in text: cell.fill=RECONSTRUCTED_FILL
                elif "unavailable" in text or text=="nan": cell.fill=UNAVAILABLE_FILL


def _narrative_frame(title:str,items:list[str])->pd.DataFrame:
    return pd.DataFrame({title:items})


def create_excel_workbook(path:Path,context:dict[str,Any],config:dict[str,Any],manifest:pd.DataFrame|None=None)->Path:
    """Create and reopen the final workbook to verify structural integrity."""
    wb=Workbook(); wb.remove(wb.active); tables=context["tables"]; skipped=[]
    mapping={
        "Intake_Audit":tables.get("intake_audit",pd.DataFrame()),"Run_Boundaries":tables.get("boundaries",pd.DataFrame()),"Steady_State":tables.get("steady",pd.DataFrame()),"VFD_Verification":tables.get("vfd",pd.DataFrame()),"Encoder_Cycle_Summary":tables.get("encoder_summary",pd.DataFrame()),"Encoder_Methods":tables.get("encoder_methods",pd.DataFrame()),"Descriptive_Statistics":tables.get("descriptive",pd.DataFrame()),"Cross_Test_Comparison":context["cross_test"],"Pressure_Response":tables.get("pressure_response",pd.DataFrame()),"Pressure_Relationships":tables.get("pressure_pairs",pd.DataFrame()),"Pressure_Phase":tables.get("pressure_phase",pd.DataFrame()),"Pressure_Consistency":tables.get("pressure_consistency",pd.DataFrame()),"Torque_Analysis":tables.get("torque",pd.DataFrame()),"Torque_Phase":tables.get("torque_phase",pd.DataFrame()),"Generator_Voltage":tables.get("generator",pd.DataFrame()),"Correlation_Regression":tables.get("correlation",pd.DataFrame()),"Signal_Quality":tables.get("quality_findings",pd.DataFrame()),"Cycle_Level_Data":tables.get("cycle_level",pd.DataFrame())}
    metadata=context["metadata"]; audit=context["audit"]
    audit={"time_source":"recorded_timestamp","time_reconstructed_flag":False,"assumed_sample_interval_s":None,"absolute_timestamp_range":f"{audit.get('timestamp_start')} to {audit.get('timestamp_end')}","duration_s":None,"sampling_jitter_available":True,"timing_dependency_note":"not applicable",**audit}
    overview=wb.create_sheet("Overview"); overview.append(["HNEI OWC Engineering Analysis","Value"])
    overview_rows=[("Input filename",metadata["input_filename"]),("Input SHA-256",metadata["input_sha256"]),("Analysis timestamp",metadata["analysis_timestamp"]),("Configuration profile",metadata["configuration_profile"]),("Worksheet analyzed",metadata.get("worksheet_analyzed") or "CSV/not applicable"),("Total records",metadata["total_records"]),("Time source",audit["time_source"]),("Time reconstructed",audit["time_reconstructed_flag"]),("Assumed sample interval (s)",audit["assumed_sample_interval_s"]),("Timestamp range",audit["absolute_timestamp_range"]),("Elapsed duration (s)",audit["duration_s"]),("Sampling jitter",("measured" if audit["sampling_jitter_available"] else "unavailable — cannot be evaluated without recorded timestamps")),("Timing-result dependency",audit["timing_dependency_note"]),("Median sampling interval (s)",audit["sampling_interval_median_s"]),("Median-derived sampling rate (Hz)",audit["median_derived_sampling_rate_hz"]),("Effective mean sampling rate (Hz)",audit["effective_mean_sampling_rate_hz"]),("Detected runs",len(context["cross_test"])),("Available channels",", ".join(metadata["available_channels"])),("Missing expected channels",", ".join(metadata["missing_expected_channels"]) or "none"),("VFD scaling status",metadata["vfd_scaling_status"]),("Output directory",metadata["output_directory"]),("Git commit",metadata["git_commit"]),("Major findings","Robust pressure response and torque increased with measured frequency; downstream sensor offsets require review."),("Highest-priority limitations","Unknown pressure/torque units, unverified pressure calibration, limited cycles, sequential run order.")]
    for row in overview_rows: overview.append(row)
    start=overview.max_row+2; overview.cell(start,1,"Test-condition summary")
    compact=context["cross_test"][[c for c in ["Run_ID","Target_Cycle_s","Target_Source","Final_Selected_Period_s","Final_Selected_Frequency_Hz","Timing_Method_Confidence","Steady_State_Sample_Count","Steady_State_Duration_s","Command_Saturation_Flag","Robust_Attenuation_Ratio","Mean_Torque","RMS_Torque","Mean_Gen_V","Primary_Quality_Finding"] if c in context["cross_test"]]]
    for c,column in enumerate(compact.columns,1): overview.cell(start+1,c,column)
    for r,row in enumerate(compact.itertuples(index=False,name=None),start+2):
        for c,value in enumerate(row,1): overview.cell(r,c,_excel_value(value))
    for cell in overview[1]: cell.fill=HEADER_FILL; cell.font=HEADER_FONT
    for cell in overview[start+1]: cell.fill=HEADER_FILL; cell.font=HEADER_FONT
    overview.freeze_panes="A2"; overview.column_dimensions["A"].width=34; overview.column_dimensions["B"].width=70
    for name in SHEET_ORDER[1:19]:
        frame=mapping[name]
        if frame.empty: skipped.append(name)
        ws=wb.create_sheet(name); _format_table(ws,frame,bool(config.get("reporting",{}).get("excel_freeze_panes",True)),bool(config.get("reporting",{}).get("excel_auto_filter",True)))
    narratives={"Methods":["See report/methods_section.md for the report-ready formal methods section.","Raw rows and measured values are preserved; inferred, reconstructed, filtered, derived, and robust values are labeled separately."],"Engineering_Interpretation":["Torque increased with measured cycle frequency and decreased with commanded cycle period.","Robust and median-cycle attenuation are primary; four-condition regressions are exploratory and correlation is not causation.","Weak-correlation phase values are not interpreted as physical phase."],"Limitations":["Pressure and torque units are unknown.","Pressure calibration and absolute differential pressure are unavailable.","Downstream pressure offset/weak response, limited cycles, sequential order, and undocumented Gen_V meaning constrain conclusions."],"Recommended_Tests":["Record targets, command voltage, and measured VFD output frequency.","Capture 10–20 steady cycles, randomize test order, verify pressure wiring/calibration and units, and add RPM/vibration measurements."],"File_Manifest":[]}
    for name in SHEET_ORDER[19:]:
        frame=manifest if name=="File_Manifest" and manifest is not None else _narrative_frame(name,narratives[name])
        ws=wb.create_sheet(name); _format_table(ws,frame)
    overview.cell(overview.max_row+2,1,"Skipped/unavailable worksheets"); overview.cell(overview.max_row,2,", ".join(skipped) if skipped else "none")
    path.parent.mkdir(parents=True,exist_ok=True); wb.save(path); check=load_workbook(path,read_only=True); assert check.sheetnames==SHEET_ORDER; check.close(); return path
